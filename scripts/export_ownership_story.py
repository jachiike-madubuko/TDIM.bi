#!/usr/bin/env python3
"""
Ownership narrative package: Setup / Conflict / Resolution.

Writes under exports/ownership_story/:
  - Bistro_Ownership_Story.xlsx
  - Bistro_Ownership_Story.pdf
  - Bistro_Ownership_Story.md
  - charts/index_with_yoy_shadow.png
  - charts/setup.png, conflict.png, resolution.png

Index month defaults to 2025-08 (= 100). Shadow series = same calendar month
prior year, also indexed to Aug 2025 sales.

Usage
-----
  python scripts/export_ownership_story.py
  python scripts/export_ownership_story.py --td-dir TD --index-month 2025-08
"""

from __future__ import annotations

import argparse
import shutil
import sys
from datetime import datetime, timezone
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Image,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "scripts"))

from export_bistro_index import (  # noqa: E402
    DEFAULT_INDEX_MONTH,
    build_index,
    month_label,
    monthly_from_supabase,
    monthly_from_td,
)

OUT_DIR = ROOT / "exports" / "ownership_story"
CHART_DIR = OUT_DIR / "charts"


def money(n: float | None) -> str:
    if n is None or (isinstance(n, float) and pd.isna(n)):
        return "—"
    sign = "-" if n < 0 else ""
    return f"{sign}${abs(n):,.0f}"


def idx_fmt(n: float | None) -> str:
    if n is None or (isinstance(n, float) and pd.isna(n)):
        return "—"
    return f"{n:.1f}"


def pct_fmt(n: float | None) -> str:
    if n is None or (isinstance(n, float) and pd.isna(n)):
        return "—"
    sign = "+" if n > 0 else ""
    return f"{sign}{n:.1f}%"


def enrich_with_shadow(indexed: pd.DataFrame) -> pd.DataFrame:
    sales_by_month = dict(zip(indexed["month"], indexed["bistro_sales"]))
    index_sales = float(indexed.loc[indexed["zone"] == "index", "index_sales"].iloc[0])

    prior_months = []
    prior_sales = []
    shadow_index = []
    yoy_delta_dollars = []
    yoy_delta_pct = []

    for month, sales in zip(indexed["month"], indexed["bistro_sales"]):
        y, m = month.split("-")
        prior = f"{int(y) - 1}-{m}"
        ps = sales_by_month.get(prior)
        prior_months.append(prior if ps is not None else None)
        prior_sales.append(round(ps, 2) if ps is not None else None)
        shadow_index.append(round(ps / index_sales * 100, 1) if ps is not None else None)
        if ps is not None and ps != 0:
            yoy_delta_dollars.append(round(sales - ps, 2))
            yoy_delta_pct.append(round((sales - ps) / abs(ps) * 100, 1))
        else:
            yoy_delta_dollars.append(None)
            yoy_delta_pct.append(None)

    out = indexed.copy()
    out["prior_year_month"] = prior_months
    out["prior_year_sales"] = prior_sales
    out["shadow_index"] = shadow_index
    out["yoy_delta_dollars"] = yoy_delta_dollars
    out["yoy_delta_pct"] = yoy_delta_pct
    return out


def build_narrative(df: pd.DataFrame, index_month: str) -> dict:
    index_row = df.loc[df["zone"] == "index"].iloc[0]
    pre = df[df["zone"] == "before"]
    post = df[df["zone"] == "after"]
    pre_avg = float(pre["index_value"].mean()) if len(pre) else None
    post_avg = float(post["index_value"].mean()) if len(post) else None
    latest = post.iloc[-1] if len(post) else None
    after_deficit = float(post["deficit_dollars"].sum()) if len(post) else 0.0

    # YoY on months that have a prior-year match (typically from index month forward)
    yoy = df[df["yoy_delta_pct"].notna()].copy()
    yoy_post = yoy[yoy["zone"].isin(["index", "after"])]
    yoy_ahead = int((yoy_post["yoy_delta_pct"] > 0).sum()) if len(yoy_post) else 0
    yoy_behind = int((yoy_post["yoy_delta_pct"] < 0).sum()) if len(yoy_post) else 0
    yoy_comp_sales = float(yoy_post["bistro_sales"].sum()) if len(yoy_post) else None
    yoy_prior_sales = float(yoy_post["prior_year_sales"].sum()) if len(yoy_post) else None
    yoy_delta = (
        (yoy_comp_sales - yoy_prior_sales) if yoy_comp_sales is not None and yoy_prior_sales else None
    )
    yoy_pct = (
        (yoy_delta / abs(yoy_prior_sales) * 100)
        if yoy_delta is not None and yoy_prior_sales
        else None
    )

    setup_title = (
        f"Climbing toward the {month_label(index_month)} mark"
        if pre_avg is not None and pre_avg < 90
        else (
            f"Approaching the {month_label(index_month)} baseline"
            if pre_avg is not None and pre_avg < 100
            else f"Already hot heading into {month_label(index_month)}"
        )
    )
    conflict_title = f"{month_label(index_month)} becomes the baseline (index = 100)"
    if post_avg is None:
        resolution_title = f"After {month_label(index_month)}"
        thesis = (
            f"{month_label(index_month)} is set as the sales baseline — "
            "load later months to see what came after."
        )
    elif post_avg < 100:
        resolution_title = f"Still below {month_label(index_month)}"
        thesis = f"Sales have not yet retaken the {month_label(index_month)} baseline."
    else:
        resolution_title = f"Above the index after {month_label(index_month)}"
        thesis = f"Sales climbed after the {month_label(index_month)} baseline and stayed above the index."

    if len(pre):
        coldest = pre.loc[pre["index_value"].idxmin()]
        hottest = pre.loc[pre["index_value"].idxmax()]
        setup_body = (
            f"The {len(pre)} month(s) before the index averaged {idx_fmt(pre_avg)} on the index scale. "
            f"Low point: {coldest['label']} at {idx_fmt(float(coldest['index_value']))}. "
            f"High point before the pivot: {hottest['label']} at {idx_fmt(float(hottest['index_value']))}. "
            f"Setup is the climb into the intentional baseline."
        )
    else:
        setup_body = "No pre-index months loaded."

    conflict_body = (
        f"{month_label(index_month)} net sales of {money(float(index_row['bistro_sales']))} "
        f"lock to index 100. This is an intentional pivot for reading everything before and after, "
        f"not a collision of series."
    )
    if pd.notna(index_row.get("yoy_delta_pct")):
        conflict_body += (
            f" Versus {index_row['prior_year_month']}, August ran "
            f"{pct_fmt(float(index_row['yoy_delta_pct']))} "
            f"({money(float(index_row['yoy_delta_dollars']))})."
        )

    if latest is not None and post_avg is not None:
        resolution_body = (
            f"After the index, {len(post)} month(s) averaged {idx_fmt(post_avg)} "
            f"({post_avg - 100:+.1f} vs 100). Latest {latest['label']} sits at "
            f"{idx_fmt(float(latest['index_value']))}. Cumulative bistro deficit versus the "
            f"August mark is {money(after_deficit)}."
        )
    else:
        resolution_body = "Load months after the index to complete the after-state."

    if yoy_pct is not None:
        yoy_note = (
            f"On {len(yoy_post)} comparable month(s) with a prior-year match "
            f"(from {yoy_post['month'].iloc[0]} through {yoy_post['month'].iloc[-1]}), "
            f"current period did {money(yoy_comp_sales)} vs {money(yoy_prior_sales)} prior "
            f"({pct_fmt(yoy_pct)}, {money(yoy_delta)}). "
            f"{yoy_ahead} month(s) ahead YoY, {yoy_behind} behind."
        )
    else:
        yoy_note = "Prior-year shadow is incomplete for this span."

    overview = [
        (
            f"{month_label(index_month)} net sales of {money(float(index_row['bistro_sales']))} "
            "are indexed to 100 — the pivot for reading everything before and after."
        ),
        f"Setup: {setup_body}",
        f"Conflict: {conflict_body}",
        f"Resolution: {resolution_body}",
        f"YoY shadow: {yoy_note}",
        "Margin and product-matrix storytelling stay offline until COGS joins on item Number.",
    ]

    return {
        "thesis": thesis,
        "setup_title": setup_title,
        "conflict_title": conflict_title,
        "resolution_title": resolution_title,
        "setup_body": setup_body,
        "conflict_body": conflict_body,
        "resolution_body": resolution_body,
        "yoy_note": yoy_note,
        "overview": overview,
        "pre_avg": pre_avg,
        "post_avg": post_avg,
        "after_deficit": after_deficit,
        "index_sales": float(index_row["bistro_sales"]),
        "index_label": month_label(index_month),
        "latest_label": str(latest["label"]) if latest is not None else None,
        "latest_index": float(latest["index_value"]) if latest is not None else None,
        "yoy_pct": yoy_pct,
        "yoy_ahead": yoy_ahead,
        "yoy_behind": yoy_behind,
        "span": f"{df['month'].iloc[0]} → {df['month'].iloc[-1]}",
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
    }


def style_index_chart(ax, title: str):
    ax.set_title(title, fontsize=12, fontweight="bold", pad=10)
    ax.set_ylabel("Index (Aug 2025 = 100)")
    ax.axhline(100, color="#059669", linestyle="--", linewidth=1.4, label="Index = 100")
    ax.grid(True, axis="y", alpha=0.25)
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.tick_params(axis="x", rotation=45, labelsize=8)


def save_charts(df: pd.DataFrame, narrative: dict) -> dict[str, Path]:
    CHART_DIR.mkdir(parents=True, exist_ok=True)
    paths: dict[str, Path] = {}

    # Main index + shadow
    fig, ax = plt.subplots(figsize=(11, 5.2), dpi=140)
    x = list(range(len(df)))
    labels = df["label"].tolist()
    ax.plot(x, df["index_value"], color="#1e293b", linewidth=2.2, marker="o", markersize=4, label="Current (vs Aug 2025)")
    shadow = df["shadow_index"]
    if shadow.notna().any():
        ax.plot(
            x,
            shadow,
            color="#94a3b8",
            linewidth=1.8,
            linestyle="--",
            marker="o",
            markersize=3.5,
            alpha=0.85,
            label="Same month last year (shadow)",
        )
    # Mark index month
    idx_i = list(df["zone"]).index("index")
    ax.scatter([idx_i], [100], s=90, color="#059669", zorder=5, label="Aug 2025 pivot")
    style_index_chart(ax, "Bistro net sales index · Aug 2025 = 100 · YoY shadow")
    ax.set_xticks(x)
    ax.set_xticklabels(labels)
    ax.legend(loc="upper right", fontsize=8, frameon=False)
    fig.tight_layout()
    paths["index"] = CHART_DIR / "index_with_yoy_shadow.png"
    fig.savefig(paths["index"], bbox_inches="tight")
    plt.close(fig)

    # Phase charts
    for zone, key, color, title in [
        ("before", "setup", "#64748b", narrative["setup_title"]),
        ("index", "conflict", "#059669", narrative["conflict_title"]),
        ("after", "resolution", "#4f46e5", narrative["resolution_title"]),
    ]:
        part = df[df["zone"] == zone]
        fig, ax = plt.subplots(figsize=(8.5, 4.2), dpi=140)
        if len(part):
            ax.bar(part["label"], part["index_value"], color=color, alpha=0.85)
            ax.axhline(100, color="#059669", linestyle="--", linewidth=1.2)
        ax.set_title(f"{key.title()}: {title}", fontsize=11, fontweight="bold")
        ax.set_ylabel("Index")
        ax.tick_params(axis="x", rotation=40, labelsize=8)
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.grid(True, axis="y", alpha=0.25)
        fig.tight_layout()
        paths[key] = CHART_DIR / f"{key}.png"
        fig.savefig(paths[key], bbox_inches="tight")
        plt.close(fig)

    return paths


def write_markdown(df: pd.DataFrame, narrative: dict, chart_paths: dict[str, Path]) -> Path:
    path = OUT_DIR / "Bistro_Ownership_Story.md"
    lines = [
        "# Bistro Ownership Story",
        "",
        f"**Thesis:** {narrative['thesis']}",
        "",
        f"_Span:_ {narrative['span']} · _Index:_ {narrative['index_label']} = 100 "
        f"({money(narrative['index_sales'])} net sales) · _Generated:_ {narrative['generated_at']}",
        "",
        "## How to read this",
        "",
        "- **Index:** every month is `sales / Aug 2025 sales × 100`. August 2025 is locked at 100.",
        "- **Shadow:** same calendar month last year, also indexed to Aug 2025. Use it to separate seasonality from the post-peak soft stretch.",
        "- **Setup / Conflict / Resolution:** before the pivot, the intentional baseline, and the after-state.",
        "",
        "## Main chart",
        "",
        f"![Index with YoY shadow]({chart_paths['index'].relative_to(OUT_DIR)})",
        "",
        "## Narrative phases",
        "",
        f"### Setup — {narrative['setup_title']}",
        "",
        narrative["setup_body"],
        "",
        f"![Setup]({chart_paths['setup'].relative_to(OUT_DIR)})",
        "",
        f"### Conflict — {narrative['conflict_title']}",
        "",
        narrative["conflict_body"],
        "",
        f"![Conflict]({chart_paths['conflict'].relative_to(OUT_DIR)})",
        "",
        f"### Resolution — {narrative['resolution_title']}",
        "",
        narrative["resolution_body"],
        "",
        f"![Resolution]({chart_paths['resolution'].relative_to(OUT_DIR)})",
        "",
        "## Ownership snapshot",
        "",
        f"| Metric | Value |",
        f"| --- | --- |",
        f"| Index month | {narrative['index_label']} ({money(narrative['index_sales'])}) |",
        f"| Pre-index avg | {idx_fmt(narrative['pre_avg'])} |",
        f"| Post-index avg | {idx_fmt(narrative['post_avg'])} |",
        f"| Latest month | {narrative['latest_label']} · {idx_fmt(narrative['latest_index'])} |",
        f"| Cumulative deficit vs index (after) | {money(narrative['after_deficit'])} |",
        f"| YoY on comparable months | {pct_fmt(narrative['yoy_pct'])} ({narrative['yoy_ahead']} ahead / {narrative['yoy_behind']} behind) |",
        "",
        "## Overview",
        "",
    ]
    for line in narrative["overview"]:
        lines.append(f"- {line}")
    lines += [
        "",
        "## Monthly detail",
        "",
        "| Month | Zone | Sales | Index | Prior-year sales | Shadow index | YoY $ | YoY % | Deficit vs index |",
        "| --- | --- | ---: | ---: | ---: | ---: | ---: | ---: | ---: |",
    ]
    for _, r in df.iterrows():
        lines.append(
            f"| {r['label']} | {r['zone']} | {money(float(r['bistro_sales']))} | "
            f"{idx_fmt(float(r['index_value']))} | "
            f"{money(float(r['prior_year_sales'])) if pd.notna(r['prior_year_sales']) else '—'} | "
            f"{idx_fmt(float(r['shadow_index'])) if pd.notna(r['shadow_index']) else '—'} | "
            f"{money(float(r['yoy_delta_dollars'])) if pd.notna(r['yoy_delta_dollars']) else '—'} | "
            f"{pct_fmt(float(r['yoy_delta_pct'])) if pd.notna(r['yoy_delta_pct']) else '—'} | "
            f"{money(float(r['deficit_dollars']))} |"
        )
    lines += [
        "",
        "## Source",
        "",
        "- Measure: Check Line Total after IPA type-in cleaning (`clean_tdim.py`).",
        "- Grain: calendar month.",
        "- Charts for slide decks: also download PNG from the Story Constructor graphs in the web app.",
        "",
    ]
    path.write_text("\n".join(lines) + "\n")
    return path


def write_excel(df: pd.DataFrame, narrative: dict) -> Path:
    path = OUT_DIR / "Bistro_Ownership_Story.xlsx"
    wb = Workbook()

    # Cover
    ws = wb.active
    ws.title = "Ownership Brief"
    header_fill = PatternFill("solid", fgColor="1E293B")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    accent = Font(bold=True, size=14, color="0F172A")

    ws["A1"] = "Bistro Ownership Story"
    ws["A1"].font = Font(bold=True, size=18, color="0F172A")
    ws["A2"] = narrative["thesis"]
    ws["A2"].font = Font(size=12, italic=True)
    ws["A3"] = f"Span {narrative['span']} · Generated {narrative['generated_at']}"

    rows = [
        ("Index month", f"{narrative['index_label']} = 100"),
        ("Index sales", money(narrative["index_sales"])),
        ("Pre-index average", idx_fmt(narrative["pre_avg"])),
        ("Post-index average", idx_fmt(narrative["post_avg"])),
        ("Latest month", f"{narrative['latest_label']} · {idx_fmt(narrative['latest_index'])}"),
        ("Cumulative deficit vs index (after)", money(narrative["after_deficit"])),
        ("YoY on comparable months", pct_fmt(narrative["yoy_pct"])),
        ("YoY months ahead / behind", f"{narrative['yoy_ahead']} / {narrative['yoy_behind']}"),
    ]
    ws["A5"] = "Snapshot"
    ws["A5"].font = accent
    for i, (k, v) in enumerate(rows, start=6):
        ws.cell(i, 1, k)
        ws.cell(i, 2, v)

    ws["A15"] = "Setup"
    ws["A15"].font = accent
    ws["A16"] = narrative["setup_title"]
    ws["A17"] = narrative["setup_body"]
    ws["A17"].alignment = Alignment(wrap_text=True)

    ws["A19"] = "Conflict"
    ws["A19"].font = accent
    ws["A20"] = narrative["conflict_title"]
    ws["A21"] = narrative["conflict_body"]
    ws["A21"].alignment = Alignment(wrap_text=True)

    ws["A23"] = "Resolution"
    ws["A23"].font = accent
    ws["A24"] = narrative["resolution_title"]
    ws["A25"] = narrative["resolution_body"]
    ws["A25"].alignment = Alignment(wrap_text=True)

    ws["A27"] = "YoY shadow"
    ws["A27"].font = accent
    ws["A28"] = narrative["yoy_note"]
    ws["A28"].alignment = Alignment(wrap_text=True)

    ws["A30"] = "Overview"
    ws["A30"].font = accent
    for i, line in enumerate(narrative["overview"], start=31):
        ws.cell(i, 1, line)
        ws.cell(i, 1).alignment = Alignment(wrap_text=True)

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 28
    ws.row_dimensions[17].height = 60
    ws.row_dimensions[21].height = 60
    ws.row_dimensions[25].height = 60
    ws.row_dimensions[28].height = 50

    # Index series sheet
    ws2 = wb.create_sheet("Index Series")
    cols = [
        "month",
        "label",
        "zone",
        "bistro_sales",
        "index_value",
        "prior_year_month",
        "prior_year_sales",
        "shadow_index",
        "yoy_delta_dollars",
        "yoy_delta_pct",
        "deficit_dollars",
        "surplus_dollars",
        "delta_vs_index_pts",
        "delta_vs_index_dollars",
    ]
    for c, name in enumerate(cols, start=1):
        cell = ws2.cell(1, c, name)
        cell.fill = header_fill
        cell.font = header_font
    for r_i, (_, row) in enumerate(df.iterrows(), start=2):
        for c, name in enumerate(cols, start=1):
            ws2.cell(r_i, c, row.get(name))
    for c in range(1, len(cols) + 1):
        ws2.column_dimensions[get_column_letter(c)].width = 18

    # Phase titles sheet
    ws3 = wb.create_sheet("Phase Titles")
    ws3["A1"] = "phase"
    ws3["B1"] = "title"
    ws3["C1"] = "body"
    for cell in (ws3["A1"], ws3["B1"], ws3["C1"]):
        cell.fill = header_fill
        cell.font = header_font
    phases = [
        ("setup", narrative["setup_title"], narrative["setup_body"]),
        ("conflict", narrative["conflict_title"], narrative["conflict_body"]),
        ("resolution", narrative["resolution_title"], narrative["resolution_body"]),
    ]
    for i, (p, t, b) in enumerate(phases, start=2):
        ws3.cell(i, 1, p)
        ws3.cell(i, 2, t)
        ws3.cell(i, 3, b)
        ws3.cell(i, 3).alignment = Alignment(wrap_text=True)
    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 48
    ws3.column_dimensions["C"].width = 80

    wb.save(path)
    return path


def write_pdf(df: pd.DataFrame, narrative: dict, chart_paths: dict[str, Path]) -> Path:
    path = OUT_DIR / "Bistro_Ownership_Story.pdf"
    styles = getSampleStyleSheet()
    title = ParagraphStyle(
        "OwnTitle",
        parent=styles["Heading1"],
        fontSize=18,
        spaceAfter=8,
        textColor=colors.HexColor("#0f172a"),
    )
    h2 = ParagraphStyle(
        "OwnH2",
        parent=styles["Heading2"],
        fontSize=13,
        spaceBefore=12,
        spaceAfter=6,
        textColor=colors.HexColor("#0f172a"),
    )
    body = ParagraphStyle(
        "OwnBody",
        parent=styles["Normal"],
        fontSize=10,
        leading=14,
        spaceAfter=6,
    )
    thesis_style = ParagraphStyle(
        "OwnThesis",
        parent=styles["Normal"],
        fontSize=11,
        leading=15,
        textColor=colors.HexColor("#334155"),
        spaceAfter=10,
    )
    meta = ParagraphStyle(
        "OwnMeta",
        parent=styles["Normal"],
        fontSize=8,
        textColor=colors.HexColor("#64748b"),
        spaceAfter=12,
    )

    doc = SimpleDocTemplate(
        str(path),
        pagesize=letter,
        leftMargin=0.7 * inch,
        rightMargin=0.7 * inch,
        topMargin=0.6 * inch,
        bottomMargin=0.6 * inch,
    )
    story = []
    story.append(Paragraph("Bistro Ownership Story", title))
    story.append(Paragraph(narrative["thesis"], thesis_style))
    story.append(
        Paragraph(
            f"Span {narrative['span']} · Index {narrative['index_label']} = 100 "
            f"({money(narrative['index_sales'])}) · Generated {narrative['generated_at']}",
            meta,
        )
    )

    snap = [
        ["Metric", "Value"],
        ["Index month", f"{narrative['index_label']} ({money(narrative['index_sales'])})"],
        ["Pre-index avg", idx_fmt(narrative["pre_avg"])],
        ["Post-index avg", idx_fmt(narrative["post_avg"])],
        ["Latest month", f"{narrative['latest_label']} · {idx_fmt(narrative['latest_index'])}"],
        ["Cumulative deficit vs index", money(narrative["after_deficit"])],
        ["YoY comparable", f"{pct_fmt(narrative['yoy_pct'])} ({narrative['yoy_ahead']} ahead / {narrative['yoy_behind']} behind)"],
    ]
    t = Table(snap, colWidths=[2.6 * inch, 4.2 * inch])
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f8fafc")),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    story.append(t)
    story.append(Spacer(1, 12))

    story.append(Paragraph("Main chart — index with YoY shadow", h2))
    story.append(Image(str(chart_paths["index"]), width=7.0 * inch, height=3.3 * inch))
    story.append(
        Paragraph(
            "Solid line: current months vs Aug 2025 = 100. Dashed shadow: same calendar month last year, also indexed to Aug 2025.",
            meta,
        )
    )

    story.append(Paragraph(f"Setup — {narrative['setup_title']}", h2))
    story.append(Paragraph(narrative["setup_body"], body))
    story.append(Image(str(chart_paths["setup"]), width=6.5 * inch, height=3.0 * inch))

    story.append(Paragraph(f"Conflict — {narrative['conflict_title']}", h2))
    story.append(Paragraph(narrative["conflict_body"], body))
    story.append(Image(str(chart_paths["conflict"]), width=6.5 * inch, height=2.6 * inch))

    story.append(Paragraph(f"Resolution — {narrative['resolution_title']}", h2))
    story.append(Paragraph(narrative["resolution_body"], body))
    story.append(Image(str(chart_paths["resolution"]), width=6.5 * inch, height=3.0 * inch))

    story.append(Paragraph("YoY shadow", h2))
    story.append(Paragraph(narrative["yoy_note"], body))

    story.append(Paragraph("Overview", h2))
    for line in narrative["overview"]:
        story.append(Paragraph(f"• {line}", body))

    # Compact table of recent comparable months
    yoy = df[df["yoy_delta_pct"].notna()].tail(12)
    if len(yoy):
        story.append(Paragraph("Comparable months (index + shadow)", h2))
        table_data = [["Month", "Index", "Shadow", "YoY %", "Sales"]]
        for _, r in yoy.iterrows():
            table_data.append(
                [
                    r["label"],
                    idx_fmt(float(r["index_value"])),
                    idx_fmt(float(r["shadow_index"])) if pd.notna(r["shadow_index"]) else "—",
                    pct_fmt(float(r["yoy_delta_pct"])) if pd.notna(r["yoy_delta_pct"]) else "—",
                    money(float(r["bistro_sales"])),
                ]
            )
        t2 = Table(table_data, colWidths=[1.3 * inch, 1.0 * inch, 1.0 * inch, 1.0 * inch, 1.3 * inch])
        t2.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
                    ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#cbd5e1")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        story.append(t2)

    story.append(Spacer(1, 10))
    story.append(
        Paragraph(
            "Source: TDIM Check Line Total after IPA cleaning. Charts also available as PNG downloads in the web Story Constructor.",
            meta,
        )
    )

    doc.build(story)
    return path


def main():
    ap = argparse.ArgumentParser(description="Export ownership Setup/Conflict/Resolution package")
    ap.add_argument("--td-dir", type=Path, default=ROOT / "TD")
    ap.add_argument("--from-supabase", action="store_true")
    ap.add_argument("--index-month", default=DEFAULT_INDEX_MONTH)
    ap.add_argument("--out-dir", type=Path, default=ROOT / "exports" / "ownership_story")
    args = ap.parse_args()

    out_dir = args.out_dir
    chart_dir = out_dir / "charts"
    out_dir.mkdir(parents=True, exist_ok=True)
    chart_dir.mkdir(parents=True, exist_ok=True)

    # Point module-level paths used by writers at this run's out dir
    global OUT_DIR, CHART_DIR
    OUT_DIR = out_dir
    CHART_DIR = chart_dir

    if args.from_supabase:
        monthly, _ = monthly_from_supabase()
    elif args.td_dir.exists():
        monthly, _ = monthly_from_td(args.td_dir)
    else:
        sys.exit(f"TD dir not found: {args.td_dir}")

    indexed = build_index(monthly, args.index_month)
    df = enrich_with_shadow(indexed)
    narrative = build_narrative(df, args.index_month)
    charts = save_charts(df, narrative)
    md = write_markdown(df, narrative, charts)
    xlsx = write_excel(df, narrative)
    pdf = write_pdf(df, narrative, charts)

    root_md = ROOT / "Bistro_Ownership_Story.md"
    # Root copies so deliverables are easy to find (exports/ is gitignored)
    root_text = md.read_text().replace(
        "](charts/",
        "](exports/ownership_story/charts/",
    )
    root_md.write_text(root_text)
    shutil.copy2(xlsx, ROOT / "Bistro_Ownership_Story.xlsx")
    shutil.copy2(pdf, ROOT / "Bistro_Ownership_Story.pdf")

    print("Wrote", md)
    print("Wrote", xlsx)
    print("Wrote", pdf)
    print("Wrote", root_md)
    for k, p in charts.items():
        print("Wrote", p)
    print(f"Thesis: {narrative['thesis']}")
    print(
        f"Index {args.index_month} = {narrative['index_sales']:,.2f} · "
        f"post avg {idx_fmt(narrative['post_avg'])} · "
        f"after deficit {money(narrative['after_deficit'])} · "
        f"YoY {pct_fmt(narrative['yoy_pct'])}"
    )


if __name__ == "__main__":
    main()
